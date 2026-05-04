"""
Utility module for reading test data from multiple sources (JSON, CSV, Excel).
Supports data-driven testing with parameterized test cases.
"""

import json
import csv
import os
from pathlib import Path


class DataReader:
    """Reads test data from JSON, CSV, or Excel files for data-driven testing."""

    BASE_DIR = Path(__file__).resolve().parent.parent
    DATA_DIR = BASE_DIR / "test_data"

    @staticmethod
    def read_json(filename: str) -> list[dict]:
        """
        Read test data from a JSON file.

        Args:
            filename: Name of the JSON file (with or without .json extension).

        Returns:
            List of dictionaries containing test data.
        """
        if not filename.endswith(".json"):
            filename += ".json"

        filepath = DataReader.DATA_DIR / filename

        if not filepath.exists():
            raise FileNotFoundError(f"Test data file not found: {filepath}")

        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        return data

    @staticmethod
    def read_csv(filename: str) -> list[dict]:
        """
        Read test data from a CSV file.

        Args:
            filename: Name of the CSV file (with or without .csv extension).

        Returns:
            List of dictionaries containing test data.
        """
        if not filename.endswith(".csv"):
            filename += ".csv"

        filepath = DataReader.DATA_DIR / filename

        if not filepath.exists():
            raise FileNotFoundError(f"Test data file not found: {filepath}")

        data = []
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(dict(row))

        return data

    @staticmethod
    def read_excel(filename: str, sheet_name: str = "Sheet1") -> list[dict]:
        """
        Read test data from an Excel file.

        Args:
            filename: Name of the Excel file (with or without .xlsx extension).
            sheet_name: Name of the worksheet to read from.

        Returns:
            List of dictionaries containing test data.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel support. Install: pip install openpyxl")

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        filepath = DataReader.DATA_DIR / filename

        if not filepath.exists():
            raise FileNotFoundError(f"Test data file not found: {filepath}")

        workbook = load_workbook(filepath)
        sheet = workbook[sheet_name]

        headers = [cell.value for cell in sheet[1]]
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for header, value in zip(headers, row):
                row_data[header] = value if value is not None else ""
            data.append(row_data)

        return data

    @staticmethod
    def get_login_test_data(source: str = "json") -> list[dict]:
        """
        Convenience method to get login test data from the specified source.

        Args:
            source: Data source type - 'json', 'csv', or 'excel'.

        Returns:
            List of dictionaries containing login test data.
        """
        readers = {
            "json": lambda: DataReader.read_json("login_data"),
            "csv": lambda: DataReader.read_csv("login_data"),
            "excel": lambda: DataReader.read_excel("login_data"),
        }

        if source not in readers:
            raise ValueError(f"Unsupported data source: {source}. Use 'json', 'csv', or 'excel'.")

        return readers[source]()

    @staticmethod
    def get_test_ids(data: list[dict]) -> list[str]:
        """
        Extract test IDs from test data for pytest parameterization.

        Args:
            data: List of test data dictionaries.

        Returns:
            List of test ID strings.
        """
        return [item.get("test_id", f"test_{i}") for i, item in enumerate(data)]

    @staticmethod
    def filter_by_type(data: list[dict], test_type: str) -> list[dict]:
        """
        Filter test data by test type (positive, negative, security).

        Args:
            data: List of test data dictionaries.
            test_type: Type to filter by.

        Returns:
            Filtered list of test data dictionaries.
        """
        return [item for item in data if item.get("test_type") == test_type]
